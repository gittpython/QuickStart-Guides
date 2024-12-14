import pandas as pd
from openpyxl import load_workbook

# Define file site_code
file_site_code = 'VGIA_SITES.xlsx'


# Function to add data to Excel file
def add_to_excel(site_code, phone_number, location):
    # Create a DataFrame from the input
    df = pd.DataFrame({'site_code': [site_code], 'phone_number': [phone_number], 'location': [location]})

    # Try to load the existing workbook, if available
    try:
        book = load_workbook(file_site_code)
        with pd.ExcelWriter(file_site_code, engine='openpyxl') as writer:
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
    except FileNotFoundError:
        # If the file does not exist, create a new one
        df.to_excel(file_site_code, index=False)


# Function to search the Excel file
def search_in_excel(column, value):
    try:
        df = pd.read_excel(file_site_code)
        result = df[df[column] == value]
        if not result.empty:
            print("Search Results:")
            print(result)
        else:
            print("No matching results found.")
    except Exception as e:
        print(f"An error occurred: {e}")


# VGIA_SITES usphone_number
add_to_excel('John Doe', 30, 'New York')
add_to_excel('Jane Smith', 25, 'Los Angeles')

search_in_excel('site_code', 'John Doe')  # Searching for a site_code in the Excel file